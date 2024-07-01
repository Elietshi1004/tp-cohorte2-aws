import psutil
import openpyxl
from datetime import datetime
import time

# Créer le classeur Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Performances du serveur"

# Ajouter les en-têtes de colonne
sheet["A1"] = "Heure"
sheet["B1"] = "Utilisation du CPU (%)"
sheet["C1"] = "Utilisation de la mémoire (%)"

# Définir la ligne de départ pour les données
row = 2

while True:
    # Récupérer les informations sur l'utilisation du CPU et de la mémoire
    cpu_percent = psutil.cpu_percent(interval=2)
    mem_percent = psutil.virtual_memory().percent

    # Obtenir l'heure actuelle
    current_time = datetime.now().strftime("%H:%M:%S")

    # Ajouter les données à la feuille Excel
    sheet.cell(row=row, column=1, value=current_time)
    sheet.cell(row=row, column=2, value=cpu_percent)
    sheet.cell(row=row, column=3, value=mem_percent)

    # Enregistrer le classeur Excel
    workbook.save("performances_serveur.xlsx")

    # Passer à la ligne suivante
    row += 1
    print("cpu percent ",cpu_percent)

    # Attendre 2 secondes avant la prochaine mesure
    time.sleep(2)
