
import openpyxl

# Ouvrir le fichier Excel
workbook = openpyxl.load_workbook("performances_serveur.xlsx")
sheet = workbook.active

# Parcourir les lignes de données
for row in range(2, sheet.max_row + 1):
    heure = sheet.cell(row=row, column=1).value
    cpu_percent = sheet.cell(row=row, column=2).value
    mem_percent = sheet.cell(row=row, column=3).value

    # Afficher les données
    print(f"Heure : {heure}, Utilisation du CPU : {cpu_percent}%, Utilisation de la mémoire : {mem_percent}%")
